# !/usr/bin/python
#  -*- coding: UTF-8 -*-
import datetime
import json
import argparse
import sys
import os
import time
import urllib.parse
import urllib
import openpyxl

today = time.strftime("%Y-%m-%d", time.localtime())


workbook = openpyxl.load_workbook('猫咪档案.xlsx')

data = workbook.active

rowNum = data.max_row  # sheet行数
colNum = data.max_column   # sheet列数


#  获取所有单元格的内容
data_list = []
for i in range(10, rowNum):
    rowlist = []
    for j in range(colNum):
        cell = data.cell(i+1, j+1).value
        if isinstance(cell, datetime.datetime):
            cell = cell.strftime("%Y-%m-%d")
            # print("a")
        if cell == None:
            cell = ''
        rowlist.append(cell)
    data_list.append(rowlist)


#  输出所有单元格的内容

rowNum -= 12

"""

for i in range(rowNum):
    for j in range(colNum):
        print(data_list[i][j], ' ',end="")
    print('\n')

"""


labels = [
    [2, 'name', lambda x:'【还没有名字】' if len(x) < 1 else x],
    [3, 'addPhotoNumber', lambda x:x],
    [4, 'nickname', lambda x:x],
    [5, 'furColor', lambda x:x],
    [6, 'classification', lambda x: '纯色' if x == 5 else '玳瑁及三花' if x ==
        4 else '奶牛' if x == 3 else '橘猫及橘白' if x == 2 else '狸花'],
    [7, 'location', lambda x:x],
    [8, 'gender', lambda x:'公' if x == 1 else '母' if x == 0 else '未知'],
    [9, 'status', lambda x:'失踪' if x == '不明' else '失踪' if x == '许久未见' else x],
    [10, 'isSterilization', lambda x:'已绝育' if x ==
     1 else '未绝育' if x == 0 else '未知/可能不适宜绝育'],
    [11, 'sterilizationTime', lambda x:str(x)],
    [12, 'birthTime', lambda x:x],
    [13, 'appearance', lambda x:x],
    [14, 'character', lambda x: '亲人可抱' if x == 6 else '亲人不可抱 可摸' if x == 5 else '薛定谔亲人' if x == 4 else '吃东西时可以一直摸' if x ==
     3 else '吃东西时可以摸一下' if x == 2 else '怕人 安全距离 1m 以内' if x == 1 else '怕人 安全距离 1m 以外' if x == 0 else '未知 数据缺失'],
    [15, 'firstSightingTime', lambda x: str(x)],
    [16, 'firstSightingLocation', lambda x: str(x)],
    [17, 'relationship', lambda x: str(x)],
    [18, 'notes', lambda x:str(x)],
    [19, 'deliveryTime', lambda x:str(x)],
    [20, 'deathTime', lambda x:x],
    [21, 'deathReason', lambda x:x],
    [22, 'audioNumber', lambda x: 0 if x == '' else str(x)],
    [23, 'moreInformation', lambda x:x],
    [24, 'missingTime', lambda x:x],

]

data_json = []

for i in range(rowNum):
    json_line = {}

    if str(data_list[i][3]) == '':
        continue

    for j in labels:
        json_line[j[1]] = j[2](data_list[i][j[0]])

    data_json.append(json_line)

names = []
for line in data_json:
    if line['addPhotoNumber'] != '':
        names.append(line['name'])


with open('all.json', 'w', encoding='utf-8') as f:
    for line in data_json:
        num = 0
        f.write('{\n')
        newLine = line.copy()

        if newLine['relationship'] != '':
            f.write('  "relatedCats": "')
            firstRela = 0
            for i in names:
                if i in newLine['relationship'] and i != newLine['name']:
                    if firstRela == 0:
                        firstRela = 1
                    else:
                        f.write(' ')
                    f.write(i)
            f.write('", \n')

        for item in line:
            if line[item] == '':
                del newLine[item]
        for item in newLine:
            num = num + 1

            a = '  "' + str(item) + '": "' + str(line[item]) + '"'
            f.write(a)
            if num < len(newLine):
                f.write(',\n')
        f.write("\n}\n")
