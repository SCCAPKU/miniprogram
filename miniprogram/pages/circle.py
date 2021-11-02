# !/usr/bin/python
#  -*- coding: UTF-8 -*-
import json
import argparse
import sys
import os
import time
import urllib.parse
import urllib
import openpyxl

today = time.strftime("%Y-%m-%d", time.localtime())


workbook = openpyxl.load_workbook('猫咪档案.xlsx')

data = workbook.active

rowNum = data.max_row  # sheet行数
colNum = data.max_column   # sheet列数


#  获取所有单元格的内容
data_list = []
for i in range(12, rowNum):
    rowlist = []
    for j in range(colNum):
        cell = data.cell(i+1, j+1).value
        if cell == None:
            cell = ''
        rowlist.append(cell)
    data_list.append(rowlist)

# print(data_list)


#  输出所有单元格的内容

rowNum -= 12

"""

for i in range(rowNum):
    for j in range(colNum):
        print(data_list[i][j], ' ',end="")
    print('\n')

"""


labels = [
    [2, '名字', lambda x:'【还没有名字】' if len(x) < 1 else x],
    [3, '是否写入图鉴', lambda x:x],
    [4, '昵称', lambda x:x],
    [5, '毛色', lambda x:x],
    # [7, '出没地点', lambda x: str(x)],
    [8, '性别', lambda x:'公' if x == 1 else '母' if x == 0 else '未知'],
    [9, '状况', lambda x:'不明' if len(x) < 1 else x],
    [10, '绝育情况', lambda x:'已绝育' if x ==
     1 else '未绝育' if x == 0 else '未知/可能不适宜绝育'],
    [11, '绝育时间', lambda x:str(x)],
    [12, '出生时间', lambda x:x],
    [14, '性格', lambda x: '亲人可抱' if x == 6 else '亲人不可抱 可摸' if x == 5 else '薛定谔亲人' if x == 4 else '吃东西时可以一直摸' if x ==
        3 else '吃东西时可以摸一下' if x == 2 else '怕人 安全距离 1m 以内' if x == 1 else '怕人 安全距离 1m 以外' if x == 0 else '未知 数据缺失'],
    [15, '第一次目击', lambda x: str(x)],
    [20, '送养时间', lambda x:str(x)],
    [21, '离世时间', lambda x:x],
    [13, '外貌', lambda x:x],
    [18, '更多', lambda x:x],
    [17, '关系', lambda x: str(x)],
    [23, '是否加音频', lambda x:x]
]

data_json = []

for i in range(rowNum):
    json_line = {}

    if len(data_list[i][5]) < 1:  # 毛色都不知道那就是没有【x
        continue

    for j in labels:
        json_line[j[1]] = j[2](data_list[i][j[0]])

    data_json.append(json_line)
    # print(json_line)

# 创建猫文件夹
if not os.path.exists('cats'):
    os.makedirs('cats')

#  用于搜索关系链接
names = []
for line in data_json:
    if line['是否写入图鉴'] != '':
        names.append(line['名字'])
# print(names)


for line in data_json:
    if line['是否写入图鉴'] != '':
        if not os.path.exists('cats/' + line['名字']):
            os.makedirs('cats/' + line['名字'])  # 创建每只猫的文件夹
        # 创建js文件
        with open('cats/' + line['名字'] + '/' + line['名字'] + '.js', 'w', encoding='utf-8') as f:
            f.write('var app = getApp()\n Page({ \n data: {\n' 'catname:"' +
                    line['名字'] + '",\n catitems:[ \n')
            for j in labels:
                # 打印名字
                if j[1] == '名字':
                    print(str(line[j[1]]))
                    continue
                if j[0] == 23:
                    continue
                if str(line[j[1]]) == '' or j[1] == '是否写入图鉴':
                    continue
                # print(j[1] + " " + str(line[j[1]]))
                f.write(
                    '{category:"' + j[1] + '",\n content:" ' + str(line[j[1]]) + '",},\n')
            f.write('\n], \nurl: app.globalData.url,\n')
            # 增加关系跳转项
            f.write('relationship:[')
            for i in names:
                if i in line['关系'] and i != line['名字']:
                    f.write('{ rela:"' + i + '"},\n')
            f.write('], \n')
            #  后面的图片数
            f.write('nums:[\n')
            for i in range(int(line['是否写入图鉴'])):
                f.write('{ ' + 'num: {} '.format(i+1) + '},\n')
            f.write('],\n')
            #  音频数
            if line['是否加音频']:
                audio = '//pku-lostangel.oss-cn-beijing.aliyuncs.com/' + \
                    line['名字']
                audio = urllib.parse.quote(audio)
                f.write('audioArr: [\n')
                for i in range(int(line['是否加音频'])):
                    f.write('{\n ' + "src: '" + audio +
                            "{}.m4a'".format(i+1) + ',\nbl: false\n},\n')
                f.write("],\n  audKey: 'https:', \n},\n")
            else:
                f.write('},')
            with open('js.txt', 'r', encoding='utf-8') as f2:
                f.write(f2.read())

        # 补充另外两个文件
        with open('cats/' + line['名字'] + '/' + line['名字'] + '.json', 'w', encoding='utf-8') as f:
            with open('json.txt', 'r', encoding='utf-8') as f2:
                f.write(f2.read())
        with open('cats/' + line['名字'] + '/' + line['名字'] + '.wxml', 'w', encoding='utf-8') as f:
            with open('wxml.txt', 'r', encoding='utf-8') as f2:
                f.write(f2.read())
                if line['名字'] == '小黄鸭':
                    with open('文案/小黄鸭.txt', 'r', encoding='utf-8') as f2:
                        f.write(f2.read())


# 几个分页的index内容（显示哪些猫）
health = []
fostered = []
dead = []
unknown = []
nainiu = []
sanhua = []
chunse = []
lihua = []
ju = []
suoyou = []


#  分类
for i in range(rowNum):
    if data_list[i][3] != '':
        if data_list[i][9] == '离世':
            dead.append((data_list[i][2], data_list[i][21]))
        if data_list[i][9] == '送养':
            fostered.append((data_list[i][2], data_list[i][20]))
        if (data_list[i][9] == '不明' or data_list[i][9] == '许久未见' or data_list[i][9] == '失踪'):
            unknown.append(data_list[i][2])
        if (data_list[i][9] == '健康' or data_list[i][9] == '口炎'):
            if data_list[i][6] == 1:
                lihua.append(data_list[i][2])
            if data_list[i][6] == 2:
                ju.append(data_list[i][2])
            if data_list[i][6] == 3:
                nainiu.append(data_list[i][2])
            if data_list[i][6] == 4:
                sanhua.append(data_list[i][2])
            if data_list[i][6] == 5:
                chunse.append(data_list[i][2])

health = lihua + ju + nainiu + sanhua + chunse
suoyou = health

# 调整寄养的时间顺序
try:
    fostered = sorted(fostered, key=lambda student: student[1], reverse=True)
except:
    print("请严格按照格式填写送养时间：2000年01月01日")

# 调整离世的时间顺序
try:
    dead = sorted(dead, key=lambda student: student[1], reverse=True)
except:
    print("请严格按照格式填写离世时间：2000年01月01日")

# 创建毛色分类的js文件
# 奶牛
if not os.path.exists('index/奶牛'):
    os.makedirs('index/' + '奶牛')  # 创建每只猫的文件夹
    # 创建js文件
with open('index/奶牛/奶牛' + '.js', 'w', encoding='utf-8') as f:
    f.write('var app = getApp()\n Page({\ndata: { \n catlist: [\n')
    for name in nainiu:
        f.write('{ name:"'+name+'"},')
    with open('js2.txt', 'r', encoding='utf-8') as f2:
        f.write(f2.read())

# 狸花
if not os.path.exists('index/狸花'):
    os.makedirs('index/' + '狸花')  # 创建每只猫的文件夹
    # 创建js文件
with open('index/狸花/狸花' + '.js', 'w', encoding='utf-8') as f:
    f.write('var app = getApp()\n Page({\ndata: { \n catlist: [\n')
    for name in lihua:
        f.write('{ name:"'+name+'"},')
    with open('js2.txt', 'r', encoding='utf-8') as f2:
        f.write(f2.read())

# 玳瑁及三花
if not os.path.exists('index/玳瑁及三花'):
    os.makedirs('index/' + '玳瑁及三花')  # 创建每只猫的文件夹
    # 创建js文件
with open('index/玳瑁及三花/玳瑁及三花' + '.js', 'w', encoding='utf-8') as f:
    f.write('var app = getApp()\n Page({\ndata: { \n catlist: [\n')
    for name in sanhua:
        f.write('{ name:"'+name+'"},')
    with open('js2.txt', 'r', encoding='utf-8') as f2:
        f.write(f2.read())

# 纯色
if not os.path.exists('index/纯色'):
    os.makedirs('index/' + '纯色')  # 创建每只猫的文件夹
    # 创建js文件
with open('index/纯色/纯色' + '.js', 'w', encoding='utf-8') as f:
    f.write('var app = getApp()\n Page({\ndata: { \n catlist: [\n')
    for name in chunse:
        f.write('{ name:"'+name+'"},')
    with open('js2.txt', 'r', encoding='utf-8') as f2:
        f.write(f2.read())

# 橘猫及橘白
if not os.path.exists('index/橘猫及橘白'):
    os.makedirs('index/' + '橘猫及橘白')  # 创建每只猫的文件夹
    # 创建js文件
with open('index/橘猫及橘白/橘猫及橘白' + '.js', 'w', encoding='utf-8') as f:
    f.write('var app = getApp()\n Page({\ndata: { \n catlist: [\n')
    for name in ju:
        f.write('{ name:"'+name+'"},')
    with open('js2.txt', 'r', encoding='utf-8') as f2:
        f.write(f2.read())

# 所有
if not os.path.exists('index/所有'):
    os.makedirs('index/' + '所有')  # 创建每只猫的文件夹
    # 创建js文件
with open('index/所有/所有' + '.js', 'w', encoding='utf-8') as f:
    f.write('var app = getApp()\n Page({\ndata: { \n catlist: [\n')
    for name in suoyou:
        f.write('{ name:"'+name+'"},')
    with open('js2.txt', 'r', encoding='utf-8') as f2:
        f.write(f2.read())

# 创建状态分类的js文件
with open('index/index' + '.js', 'w', encoding='utf-8') as f:
    f.write('var app = getApp()\n Page({\ndata: { \n')
    #  fostered
    f.write(' fostered_catlist: [\n')
    for name in fostered:
        f.write('{ name:"'+name[0]+'"},\n')
    f.write('],\n')
    #  unknown
    f.write(' unknown_catlist: [\n')
    for name in unknown:
        f.write('{ name:"'+name+'"},\n')
    f.write('],\n')
    #  dead
    f.write(' dead_catlist: [\n')
    for name in dead:
        f.write('{ name:"'+name[0]+'"},\n')
    f.write('],\n')
    with open('js_index.txt', 'r', encoding='utf-8') as f2:
        f.write(f2.read())


# print(fostered)
